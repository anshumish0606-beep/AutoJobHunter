import json
import os
from datetime import datetime
from pathlib import Path


class ReportGenerator:
    """
    Generates beautiful HTML and Excel reports from job results.
    """

    def __init__(self, output_dir: str = "output/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_html_report(self, all_jobs: list) -> str:
        """Generate a beautiful HTML report."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"jobs_report_{date_str}.html"
        filepath = self.output_dir / filename

        # Group by portal
        by_portal = {}
        for job in all_jobs:
            portal = job.get("portal", "Unknown")
            if portal not in by_portal:
                by_portal[portal] = []
            by_portal[portal].append(job)

        # Count by relevance
        high = sum(1 for j in all_jobs if j.get("relevance") == "high")
        medium = sum(1 for j in all_jobs if j.get("relevance") == "medium")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>AutoJobHunter Report — {timestamp}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; }}
  header {{ background: linear-gradient(135deg, #1a73e8, #0d47a1); color: white; padding: 30px 40px; }}
  header h1 {{ font-size: 28px; }}
  header p {{ opacity: 0.85; margin-top: 5px; }}
  .stats {{ display: flex; gap: 20px; padding: 20px 40px; background: white; border-bottom: 1px solid #e0e0e0; }}
  .stat-card {{ background: #f8f9fa; border-radius: 10px; padding: 15px 25px; text-align: center; min-width: 120px; }}
  .stat-card .num {{ font-size: 32px; font-weight: bold; color: #1a73e8; }}
  .stat-card .label {{ font-size: 13px; color: #666; margin-top: 4px; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 20px 40px; }}
  .portal-section {{ margin-bottom: 40px; }}
  .portal-title {{ font-size: 20px; font-weight: 600; padding: 12px 0; border-bottom: 2px solid #1a73e8; margin-bottom: 15px; color: #1a73e8; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  th {{ background: #1a73e8; color: white; padding: 12px 15px; text-align: left; font-size: 13px; }}
  td {{ padding: 12px 15px; border-bottom: 1px solid #f0f0f0; font-size: 13px; vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover {{ background: #f8f9ff; }}
  .badge {{ display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 11px; font-weight: 600; }}
  .badge-high {{ background: #e8f5e9; color: #2e7d32; }}
  .badge-medium {{ background: #fff8e1; color: #f57f17; }}
  .badge-low {{ background: #fce4ec; color: #c62828; }}
  .apply-btn {{ display: inline-block; background: #1a73e8; color: white; padding: 6px 14px; border-radius: 20px; text-decoration: none; font-size: 12px; font-weight: 600; }}
  .apply-btn:hover {{ background: #1557b0; }}
  .no-link {{ color: #999; font-size: 12px; }}
  footer {{ text-align: center; padding: 20px; color: #999; font-size: 12px; }}
</style>
</head>
<body>

<header>
  <h1>🤖 AutoJobHunter Report</h1>
  <p>Generated on {timestamp} | Auto-searched across {len(by_portal)} portals</p>
</header>

<div class="stats">
  <div class="stat-card"><div class="num">{len(all_jobs)}</div><div class="label">Total Jobs</div></div>
  <div class="stat-card"><div class="num">{high}</div><div class="label">High Match</div></div>
  <div class="stat-card"><div class="num">{medium}</div><div class="label">Medium Match</div></div>
  <div class="stat-card"><div class="num">{len(by_portal)}</div><div class="label">Portals Searched</div></div>
</div>

<div class="container">
"""

        for portal, jobs in by_portal.items():
            html += f"""
  <div class="portal-section">
    <div class="portal-title">📌 {portal} — {len(jobs)} jobs found</div>
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>Job Title</th>
          <th>Company</th>
          <th>Location</th>
          <th>Experience</th>
          <th>Salary</th>
          <th>Posted</th>
          <th>Match</th>
          <th>Apply</th>
        </tr>
      </thead>
      <tbody>
"""
            for i, job in enumerate(jobs, 1):
                relevance = job.get("relevance", "medium")
                badge_class = f"badge-{relevance}"
                link = job.get("apply_link", "")
                apply_html = f'<a href="{link}" target="_blank" class="apply-btn">Apply →</a>' if link else '<span class="no-link">No link</span>'

                html += f"""
        <tr>
          <td>{i}</td>
          <td><strong>{job.get('title', 'N/A')}</strong></td>
          <td>{job.get('company', 'N/A')}</td>
          <td>{job.get('location', 'N/A')}</td>
          <td>{job.get('experience', 'Fresher') or 'Fresher'}</td>
          <td>{job.get('salary', 'Not disclosed') or 'Not disclosed'}</td>
          <td>{job.get('date_posted', 'Recent') or 'Recent'}</td>
          <td><span class="badge {badge_class}">{relevance.upper()}</span></td>
          <td>{apply_html}</td>
        </tr>
"""
            html += """
      </tbody>
    </table>
  </div>
"""

        html += f"""
</div>
<footer>AutoJobHunter — Automated Job Search System | {timestamp}</footer>
</body>
</html>"""

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

        print(f"✅ HTML Report saved: {filepath}")
        return str(filepath)

    def generate_excel_report(self, all_jobs: list) -> str:
        """Generate Excel report with openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("⚠️ openpyxl not installed, skipping Excel report")
            return ""

        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"jobs_report_{date_str}.xlsx"
        filepath = self.output_dir / filename

        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        headers = ["#", "Job Title", "Company", "Location", "Experience", "Salary", "Date Posted", "Portal", "Relevance", "Apply Link"]
        header_fill = PatternFill("solid", fgColor="1a73e8")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, job in enumerate(all_jobs, 2):
            ws_summary.cell(row=row, column=1, value=row - 1)
            ws_summary.cell(row=row, column=2, value=job.get("title", ""))
            ws_summary.cell(row=row, column=3, value=job.get("company", ""))
            ws_summary.cell(row=row, column=4, value=job.get("location", ""))
            ws_summary.cell(row=row, column=5, value=job.get("experience", "Fresher"))
            ws_summary.cell(row=row, column=6, value=job.get("salary", "Not disclosed"))
            ws_summary.cell(row=row, column=7, value=job.get("date_posted", "Recent"))
            ws_summary.cell(row=row, column=8, value=job.get("portal", ""))
            ws_summary.cell(row=row, column=9, value=job.get("relevance", "medium").upper())
            link_cell = ws_summary.cell(row=row, column=10, value=job.get("apply_link", ""))
            if job.get("apply_link"):
                link_cell.hyperlink = job["apply_link"]
                link_cell.font = Font(color="1a73e8", underline="single")

            # Color by relevance
            if job.get("relevance") == "high":
                for col in range(1, 11):
                    ws_summary.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E8F5E9")
            elif job.get("relevance") == "medium":
                for col in range(1, 11):
                    ws_summary.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF8E1")

        # Auto-fit columns
        col_widths = [5, 35, 25, 20, 15, 20, 15, 12, 12, 50]
        for col, width in enumerate(col_widths, 1):
            ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        wb.save(filepath)
        print(f"✅ Excel Report saved: {filepath}")
        return str(filepath)

    def generate_json_report(self, all_jobs: list) -> str:
        """Save raw data as JSON."""
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"jobs_data_{date_str}.json"
        filepath = self.output_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({
                "generated_at": datetime.now().isoformat(),
                "total_jobs": len(all_jobs),
                "jobs": all_jobs
            }, f, indent=2, ensure_ascii=False)

        return str(filepath)
